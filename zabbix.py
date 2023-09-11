import requests
import json
import locale
from datetime import date
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill, Font

# Local Functions.
from functions.getDateInfo import getDateInfo

zabbix_api_url = 'http://zabbix.sptc.policiacientifica.net/zabbix/api_jsonrpc.php'
zabbix_user = 'raul.rlbl'
zabbix_password = 'R@ul1605'

item_keys = ['system.cpu.util', 'vm.memory.util', 'vfs.fs.size[C:,pused]']

session = requests.Session()
response = session.post(zabbix_api_url, json={
    'jsonrpc': '2.0',
    'method': 'user.login',
    'params': {
        'user': zabbix_user,
        'password': zabbix_password
    },
    'id': 1
})
auth_token = response.json()['result']

groups = [
    {
        "id":"33",
        "name":"Prioridade 1",
        "xlsx": "prioridade1.xlsx"
    },
    {
        "id":"34",
        "name":"Prioridade 2",
        "xlsx": "prioridade2.xlsx"
    },
]

for index, prioridade in enumerate(groups):
    try:
        wb = load_workbook(prioridade['xlsx'])
    except FileNotFoundError:
        wb = Workbook()
        wb.remove(wb['Sheet'])

    wb.create_sheet(getDateInfo.dateFormated())

    ws = wb[getDateInfo.dateFormated()]
    ws.append(['HOSTNAME', 'CPU USAGE', 'MEMORY USAGE', 'DISK USAGE'])

    response = session.post(zabbix_api_url, json={
        'jsonrpc': '2.0',
        'method': 'host.get',
        'params': {
            'output': ['hostid', 'host'],
            'groupids': groups[index]['id'],
            'sortfield': 'host'
            },
            'auth': auth_token,
            'id': 3
    })

    hosts = response.json()['result']

    data = []

    for host in hosts:
        dataPerHost = []
        dataPerHost.append(host['host'])
        for item_key in item_keys:
            response = session.post(zabbix_api_url, json={
            'jsonrpc': '2.0',
            'method': 'item.get',
            'params': {
                'output': ['itemid', 'name'],
                'hostids': host['hostid'],
                'sortfield': 'name',
                'search': {
                    "key_": item_key
                },
            },
            'auth': auth_token,
            'id': 3
            })

            items = response.json()['result']

            for item in items:
                response = session.post(zabbix_api_url, json={
                'jsonrpc': '2.0',
                'method': 'history.get',
                'params': {
                    'output': 'extend',
                    'history': 0,  # 0 para valores numéricos
                    'itemids': item['itemid'],
                    'sortfield': 'clock',
                    'sortorder': 'DESC',
                    'limit': 1
                },
                'auth': auth_token,
                'id': 4
            })

            if (response.json()['result']):
                value = response.json()['result'][0]['value']
                dataPerHost.append(f"{float(value):.2f}%")
            else:
                dataPerHost.append("Collector not working.")

        data.append(dataPerHost)
        print(dataPerHost)
    for item in data:
        ws.append(item)

    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name='Calibri Light', size=11)

    for cell in ws[1]:
        cell.font = Font(name="Calibri Light",bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    wb.save(f"{groups[index]['name']} - {getDateInfo.nameMonth()}.xlsx")