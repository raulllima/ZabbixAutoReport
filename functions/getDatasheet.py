import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill, Font

class getDatasheet():
    def __init__(self, pathFile, excelFile) -> None:
        self.style = self.Style()

        self.pathFile = pathFile
        self.excelFile= excelFile
    
        try:
            self.wb = load_workbook(pathFile + excelFile)
        except FileNotFoundError:
            self.wb = Workbook()
            self.wb.remove(self.wb['Sheet'])

    def getRows(self):
        self.ws.iter_rows()

    def addSheet(self, nameSheet):
        self.wb.create_sheet(nameSheet)
        self.ws = self.wb[nameSheet]

    def addRow(self, data, styleFill=None):
        self.ws.append(data)

        if styleFill:
            for column, item in enumerate(data, start=1):
                if column != 1:
                    currentLine = self.ws.max_row
                    itemStrFormatted = re.findall(r'(\d+\.\d+)%', item)
                
                    if itemStrFormatted:
                        cell = self.ws.cell(row=currentLine, column=column)
                        itemFloatFormatted = float(itemStrFormatted[0])
                        
                        for index, color in enumerate(styleFill['setScaleColor']):
                            if itemFloatFormatted >= color['startFilter'] and itemFloatFormatted < color['endFilter']:
                                if color['font']:
                                    cell.font = Font(**color['font'])
                                
                                if color['fill']:
                                    cell.fill = PatternFill(**color['fill'])
                    else:
                        cell = self.ws.cell(row=currentLine, column=column)
                        cell.font = Font(size=12, bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="434343", end_color="434343", fill_type="solid")

    def saveFile(self) -> None:
        self.wb.save(self.pathFile + self.excelFile) 

    class Style:
        def font(self, objParams):
            Font(name='Calibri Light', size=11)
        
        def fill(self, objParams):
            PatternFill(**objParams)

        def alignment(self, objParams):
            Alignment(**objParams)