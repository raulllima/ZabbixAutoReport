from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill, Font

wb = Workbook()
ws = wb.active

header = ['Hostname','CPU', 'Memory', 'Disk']

data0 = ['NAME1', '21.98', '73.32', '100.00']
data1 = ['NAME2', '40.74', '19.11', '27.88']

ws.append(header)

allData = [data0, data1]

for data in allData:
    ws.append(data)

    for coluna, item in enumerate(data, start=1):
        if coluna != 1:  # Ignora a coluna de "Hostname"
            try:
                valor_float = float(item)
                if valor_float > 70:
                    rawValue = str(item)
                    linha_atual = ws.max_row

                    cell = ws.cell(row=linha_atual, column=coluna)
                    cell.font = Font(size=12, bold=True, color="FF0000")
                    cell.fill = PatternFill(start_color="FFC8C8", end_color="FFC8C8", fill_type="solid")

            except ValueError:
                pass


wb.save('exemplo.xlsx')
wb.close()